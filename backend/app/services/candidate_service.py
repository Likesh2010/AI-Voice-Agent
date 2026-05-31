import csv
import io
import re
from typing import List, Dict, Any, Tuple, Optional
from sqlalchemy.orm import Session
from ..repositories.repos import candidate_repo, campaign_repo
from ..models import Candidate

# Email regex validation
EMAIL_REGEX = re.compile(r"[^@]+@[^@]+\.[^@]+")


class CandidateService:
    @staticmethod
    def validate_candidate_data(data: Dict[str, str]) -> Tuple[bool, Optional[str]]:
        """Validate candidate fields. Returns (is_valid, error_reason)."""
        name = data.get("name", "").strip()
        email = data.get("email", "").strip()
        phone = data.get("phone_number", "").strip() or data.get("phone", "").strip()

        if not name:
            return False, "Candidate name is required"
        if not email:
            return False, "Candidate email is required"
        if not EMAIL_REGEX.match(email):
            return False, f"Invalid email format: {email}"
        if not phone:
            return False, "Candidate phone number is required"
        
        return True, None

    @classmethod
    def parse_csv(cls, file_content: str) -> List[Dict[str, str]]:
        """Parse raw CSV text into list of dicts."""
        candidates = []
        # Support both comma and semicolon separators
        dialect = 'excel'
        if ';' in file_content.split('\n')[0]:
            dialect = 'excel-tab' # fallback/check
        
        reader = csv.DictReader(io.StringIO(file_content))
        
        # Normalize keys to match lowercase and underscores
        for row in reader:
            normalized_row = {}
            for k, v in row.items():
                if not k:
                    continue
                clean_key = k.strip().lower().replace(" ", "_")
                normalized_row[clean_key] = v.strip() if v else ""
            
            # Map common synonyms
            if "name" not in normalized_row:
                # search for full_name or candidate_name
                for key in ["full_name", "candidate_name"]:
                    if key in normalized_row:
                        normalized_row["name"] = normalized_row[key]
            
            if "phone_number" not in normalized_row:
                for key in ["phone", "telephone", "mobile", "contact"]:
                    if key in normalized_row:
                        normalized_row["phone_number"] = normalized_row[key]

            candidates.append(normalized_row)
            
        return candidates

    @classmethod
    def parse_excel(cls, file_bytes: bytes) -> List[Dict[str, str]]:
        """Parse Excel file content into list of dicts dynamically using openpyxl."""
        candidates = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
            sheet = wb.active
            
            headers = []
            first_row = True
            for row in sheet.iter_rows(values_only=True):
                if first_row:
                    headers = [str(cell).strip().lower().replace(" ", "_") if cell else "" for cell in row]
                    first_row = False
                    continue
                
                if not any(row):  # skip empty lines
                    continue
                
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_data[headers[idx]] = str(cell).strip() if cell is not None else ""
                
                # Map synonyms
                if "name" not in row_data:
                    for key in ["full_name", "candidate_name"]:
                        if key in row_data:
                            row_data["name"] = row_data[key]
                if "phone_number" not in row_data:
                    for key in ["phone", "telephone", "mobile", "contact"]:
                        if key in row_data:
                            row_data["phone_number"] = row_data[key]
                            
                candidates.append(row_data)
        except ImportError:
            # Fallback message if openpyxl is not installed (should be in requirements)
            raise ValueError("openpyxl package is required to parse Excel spreadsheets. Please run pip install openpyxl")
        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")

        return candidates

    @classmethod
    def import_candidates(
        cls, db: Session, campaign_id: int, candidate_list: List[Dict[str, str]]
    ) -> Dict[str, Any]:
        """Validate, deduplicate, import candidates and link them to a campaign."""
        total_parsed = len(candidate_list)
        valid_count = 0
        duplicates_skipped = 0
        inserted_count = 0
        imported_candidates = []

        for item in candidate_list:
            is_valid, _ = cls.validate_candidate_data(item)
            if not is_valid:
                continue
            
            valid_count += 1
            email = item.get("email").strip()
            name = item.get("name").strip()
            phone = item.get("phone_number").strip() or item.get("phone", "").strip()
            experience = item.get("experience", "").strip()
            skills = item.get("skills", "").strip()
            location = item.get("location", "").strip()

            # 1. Check if candidate already exists in the global DB
            candidate = candidate_repo.get_by_email(db, email)
            if candidate:
                duplicates_skipped += 1
                # Update basic candidate profile details if they are empty
                updated = False
                if not candidate.experience and experience:
                    candidate.experience = experience
                    updated = True
                if not candidate.skills and skills:
                    candidate.skills = skills
                    updated = True
                if not candidate.location and location:
                    candidate.location = location
                    updated = True
                if updated:
                    db.add(candidate)
                    db.commit()
            else:
                # Create a new Candidate record
                candidate = Candidate(
                    name=name,
                    email=email,
                    phone_number=phone,
                    phone=phone, # compatibility
                    experience=experience,
                    skills=skills,
                    location=location
                )
                candidate = candidate_repo.create(db, obj_in=candidate)
                inserted_count += 1

            # 2. Link Candidate to the Campaign
            candidate_repo.link_candidate_to_campaign(db, candidate.id, campaign_id)
            imported_candidates.append(candidate)

        return {
            "summary": {
                "total_parsed": total_parsed,
                "valid": valid_count,
                "duplicates_skipped": duplicates_skipped,
                "inserted": inserted_count
            },
            "candidates": imported_candidates
        }


candidate_service = CandidateService()
